import openpyxl

def actualizar_columnas(ruta_archivo):
    print(f"Cargando el archivo: {ruta_archivo}")
    # openpyxl.load_workbook no reescribe la hoja entera, solo carga la existente
    # preservando estilos, formatos y otra información (a diferencia de pandas).
    wb = openpyxl.load_workbook(ruta_archivo)
    sheet = wb.active
    
    # Mapeo de índices de columnas (basado en la fila 1)
    # openpyxl es 1-indexed, pero usaremos el índice 0-indexed para la lógica y sumaremos 1
    headers = {cell.value: idx + 1 for idx, cell in enumerate(sheet[1])}
    
    col_arq = headers.get("Arquitectura")
    col_audio = headers.get("audio")
    col_image = headers.get("image")
    col_video = headers.get("video") # Por si también deseas incluir video
    
    if not col_arq:
        print("No se encontró la columna 'Arquitectura'.")
        return

    print("Procesando las filas...")
    filas_modificadas = 0
    for row in range(2, sheet.max_row + 1):
        arquitectura = sheet.cell(row=row, column=col_arq).value
        
        if arquitectura and isinstance(arquitectura, str):
            arq_lower = arquitectura.lower()
            
            # Popula las columnas si existen en el archivo
            if col_audio:
                sheet.cell(row=row, column=col_audio).value = "Sí" if "audio" in arq_lower else "No"
            
            if col_image:
                sheet.cell(row=row, column=col_image).value = "Sí" if "image" in arq_lower else "No"
                
            if col_video:
                sheet.cell(row=row, column=col_video).value = "Sí" if "video" in arq_lower else "No"
                
            filas_modificadas += 1

    print(f"Se procesaron {filas_modificadas} filas. Guardando cambios en el archivo...")
    # Guardar sobreescribe solo las celdas editadas, manteniendo el archivo original intacto
    wb.save(ruta_archivo)
    print("¡Proceso completado exitosamente!")

if __name__ == "__main__":
    archivo = r"C:\source\N8N\Modelos_OpenRouter_Final.xlsx"
    actualizar_columnas(archivo)
