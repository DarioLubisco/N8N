import openpyxl

# Puntajes estimados basados en discusiones de foros especializados (Nous Research, Reddit LocalLLaMA, GitHub)
# sobre la capacidad de cada modelo para actuar como Orquestador Principal en el framework de Hermes Agent.
# Los modelos de Nous (como Hermes 4) y los modelos grandes de razonamiento profundo (Claude Sonnet/Opus, DeepSeek V4)
# obtienen las mejores puntuaciones para mantener el "closed learning loop" y la persistencia de memoria.
PUNTAJES_HERMES = {
    'DeepSeek: R1 Distill Qwen 32B': 88,
    'Qwen: Qwen Plus 0728 (thinking)': 89,
    'DeepSeek: R1 Distill Llama 70B': 92,
    'Anthropic: Claude Opus 4.7': 98,
    'DeepSeek: DeepSeek V4 Flash': 95,
    'Qwen2.5 72B Instruct': 93,
    'DeepSeek: DeepSeek V4 Pro': 97,
    'Anthropic: Claude Sonnet 4.6': 99, # Considerado el rey actual de orquestación de agentes
    'Meta: Llama 3.2 11B Vision Instruct': 75,
    'Google: Gemini 3.1 Pro Preview Custom Tools': 96,
    'Qwen: Qwen3 VL 32B Instruct': 85,
    'Google: Gemini 2.5 Flash': 90,
    'Mistral: Ministral 3 8B 2512': 70,
    'Meta: Llama 3.3 70B Instruct': 94,
    'Mistral: Ministral 3 14B 2512': 76,
    'Google: Gemini 2.5 Flash Lite Preview 09-2025': 80,
    'Google: Gemini 2.5 Flash Lite': 82,
    'OpenAI: GPT-4o-mini (2024-07-18)': 87,
    'OpenAI: GPT-4o-mini': 88,
    'NVIDIA: Nemotron Nano 12B 2 VL': 72,
    'Google: Gemini 3.1 Flash Lite Preview': 85,
    'Nous: Hermes 4 70B': 96 # Modelo nativo de Nous Research, altamente optimizado para Hermes
}

def agregar_columna_hermes(ruta_archivo):
    print(f"Cargando el archivo: {ruta_archivo}")
    wb = openpyxl.load_workbook(ruta_archivo)
    sheet = wb.active
    
    headers = {cell.value: idx + 1 for idx, cell in enumerate(sheet[1])}
    
    col_nombre = headers.get("Nombre")
    col_estado = headers.get("Estado")
    
    if not col_nombre or not col_estado:
        print("No se encontraron las columnas 'Nombre' o 'Estado'.")
        return

    # Validar si existe la columna, si no, crearla al final
    nombre_columna_nueva = "Puntaje Orquestador Hermes"
    
    if nombre_columna_nueva in headers:
        col_orquestador = headers[nombre_columna_nueva]
    else:
        col_orquestador = sheet.max_column + 1
        sheet.cell(row=1, column=col_orquestador).value = nombre_columna_nueva

    print("Actualizando puntajes específicos de orquestador Hermes para modelos activos...")
    filas_modificadas = 0
    
    for row in range(2, sheet.max_row + 1):
        estado = str(sheet.cell(row=row, column=col_estado).value).strip().lower()
        
        if estado == "activo":
            nombre_modelo = sheet.cell(row=row, column=col_nombre).value
            
            puntaje = PUNTAJES_HERMES.get(nombre_modelo)
            if puntaje is not None:
                sheet.cell(row=row, column=col_orquestador).value = puntaje
                filas_modificadas += 1

    print(f"Se actualizaron {filas_modificadas} modelos con su puntaje de Orquestador Hermes.")
    print("Guardando archivo...")
    wb.save(ruta_archivo)
    print("¡Archivo guardado intacto con la nueva columna!")

if __name__ == "__main__":
    archivo = r"C:\source\N8N\Modelos_OpenRouter_Final.xlsx"
    agregar_columna_hermes(archivo)
